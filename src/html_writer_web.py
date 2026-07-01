"""Excel チェックシート → 自己完結 HTML ビューワー生成.

使い方:
    from html_writer_web import write_html
    write_html(sheet_data, out_path)

sheet_data は excel_writer.py が受け取る同じ辞書構造を想定しているが、
既存の Excel ファイルからも読み込める (read_excel_to_html を参照)。
"""
from __future__ import annotations

import json
import os
import re
from datetime import datetime


def _translate_cell_val(val: str, i18n: dict) -> str:
    if not i18n.get("value_map") and not i18n.get("prefix_map"):
        return val
    vm = i18n.get("value_map", {})

    # exact match
    if val in vm:
        return vm[val]

    # dot-separated compound (出張実態・労務・金額規程 etc.)
    if "・" in val and not any(c in val for c in ("(", "/")):
        parts = val.split("・")
        if all(p in vm for p in parts):
            return " • ".join(vm[p] for p in parts)

    # prefix map
    for ja_pfx, uz_pfx in i18n.get("prefix_map", []):
        if val.startswith(ja_pfx):
            return uz_pfx + val[len(ja_pfx):]

    # dynamic regex patterns
    m = re.match(r'対象件数\(伝票\)=(\d+) / 未承認件数=(\d+) / 承認済=(\d+)', val)
    if m:
        return f"Hujjat soni={m.group(1)} / Tasdiqlanmagan={m.group(2)} / Tasdiqlangan={m.group(3)}"

    m = re.match(r'勤怠日欠落 \(明細(.+)\)', val)
    if m:
        return f"Davomati yo'q (qator {m.group(1)})"

    m = re.match(r'深夜移動 要確認 \(明細(.+)\)', val)
    if m:
        return f"Kech tun harakati tekshiruvi (qator {m.group(1)})"

    m = re.match(r'非労働日/休暇に移動 \(明細(.+)\)', val)
    if m:
        return f"Dam olish kuni harakat (qator {m.group(1)})"

    m = re.match(r'金額上限超過 \((\d+)件\)', val)
    if m:
        return f"Summa limiti oshdi ({m.group(1)} ta)"

    m = re.match(r'金額上限超過 \((\d+)件\) / 要確認 \((\d+)件\)', val)
    if m:
        return f"Summa limiti oshdi ({m.group(1)} ta) / Tekshirish kerak ({m.group(2)} ta)"

    m = re.match(r'二重申請疑い: 強一致の重複明細が (\d+) 件。', val)
    if m:
        return f"Ikki marta ariza gumonasi: {m.group(1)} ta takror qator topildi"

    m = re.match(r'勤怠データ欠落: (\S+) 月分未提供', val)
    if m:
        return f"Davomati yo'q: {m.group(1)} oy taqdim etilmagan"

    # 伝票No.XXXX: 出張日が特定できません (no date range)
    m = re.match(r'伝票No\.([^:]+): 出張日\(移動/到着明細\)が特定できません。$', val)
    if m:
        return f"Hujjat No.{m.group(1)}: Safari sanasi (harakat/borish qatorlari) aniqlanmadi."

    # 伝票No.XXXX(PERIOD): MM oy davomati yo'q → borish nuqtasi topilmadi
    m = re.match(
        r'伝票No\.([^(]+)\(([^)]+)\): (.+)の出勤簿が未提供のため出張実態は暫定評価です。'
        r'到着地が顧客マスタと突合できないため、訪問先の妥当性を別途ご確認ください。$', val)
    if m:
        return (
            f"Hujjat No.{m.group(1)}({m.group(2)}): {m.group(3)} oy davomati taqdim etilmagan — "
            f"safari holati vaqtinchalik baholandi. "
            f"Borish nuqtasi mijozlar bazasida topilmadi, tashrif maqsadini alohida tekshiring."
        )

    # 伝票No.XXXX(PERIOD): MM oy davomati yo'q → davomati kelib tushgach
    m = re.match(
        r'伝票No\.([^(]+)\(([^)]+)\): (.+)の出勤簿が未提供のため出張実態は暫定評価です。'
        r'出勤簿提供後に最終確認をお願いします。$', val)
    if m:
        return (
            f"Hujjat No.{m.group(1)}({m.group(2)}): {m.group(3)} oy davomati taqdim etilmagan — "
            f"safari holati vaqtinchalik baholandi. Davomati kelib tushgach yakuniy tekshiruv."
        )

    # 伝票No.XXXX(PERIOD): DAYS ta'til/ishdan qolish kuni
    m = re.match(
        r'伝票No\.([^(]+)\(([^)]+)\): (.+)は勤怠上 休暇/欠勤 です。'
        r'出張実態と整合しないため、日付の誤りか休暇申請の取消をご確認ください。$', val)
    if m:
        return (
            f"Hujjat No.{m.group(1)}({m.group(2)}): {m.group(3)} — "
            f"davomat bo'yicha ta'til/ishdan qolish kuni. "
            f"Safari holati bilan mos kelmaydi; sana xatosi yoki ta'til bekor qilishni tekshiring."
        )

    # 伝票No.XXXX(PERIOD): DAYS davomat kiritilmagan
    m = re.match(
        r'伝票No\.([^(]+)\(([^)]+)\): (.+)の勤怠が未入力です。出勤簿の入力状況をご確認ください。$',
        val)
    if m:
        return (
            f"Hujjat No.{m.group(1)}({m.group(2)}): {m.group(3)} — "
            f"davomat kiritilmagan. Davomat varaqasi to'ldirilganligini tekshiring."
        )

    # 伝票No.XXXX(PERIOD): borish nuqtasi mijozlar bazasida topilmadi
    m = re.match(
        r'伝票No\.([^(]+)\(([^)]+)\): 到着地が顧客マスタと突合できません。'
        r'訪問先\(取引先\)名の表記をご確認ください。$', val)
    if m:
        return (
            f"Hujjat No.{m.group(1)}({m.group(2)}): "
            f"Borish nuqtasi mijozlar bazasida topilmadi. "
            f"Tashrif nomi (hamkor) yozilishini tekshiring."
        )

    # 伝票No.XXXX(PERIOD): safari holati etarli dalil yo'q
    m = re.match(
        r'伝票No\.([^(]+)\(([^)]+)\): 出張実態の裏付けが不足しています。'
        r'出勤簿および訪問先をご確認ください。$', val)
    if m:
        return (
            f"Hujjat No.{m.group(1)}({m.group(2)}): "
            f"Safari holati uchun etarli dalil yo'q. Davomat va tashrif joyini tekshiring."
        )

    # 想定承認者 NAME は確認のみ (approval_route dynamic)
    m = re.match(
        r'想定承認者 (.+) は確認のみのため正式承認に該当しません。正規の出張命令者による承認をご確認ください。$',
        val)
    if m:
        return (
            f"Tasdiqlash: {m.group(1)} faqat 'tekshirish' rolidadir — rasmiy tasdiqlash hisoblanmaydi. "
            f"To'g'ri safari buyrug'i beruvchi tomonidan tasdiqlashni tekshiring."
        )

    # 想定の出張命令者 NAME による承認が確認できません (approval_route dynamic)
    m = re.match(
        r'想定の出張命令者 (.+) による承認が確認できません。承認ルートをご確認ください。$', val)
    if m:
        return (
            f"Safari buyrug'i beruvchi {m.group(1)} tomonidan tasdiqlash topilmadi. "
            f"Tasdiqlash zanjirini tekshiring."
        )

    # compound with semicolons (non-labour + missing)
    if "; " in val:
        parts = val.split("; ")
        translated = [_translate_cell_val(p, i18n) for p in parts]
        return "; ".join(translated)

    return val


# ── ステータスラベル → CSS クラス ──────────────────────────────────────────
def _status_class(val: str) -> str:
    v = str(val).strip()
    if v == "NG":
        return "s-ng"
    if v in ("要確認", "要確認(勤怠データ欠落)"):
        return "s-warn"
    if v == "OK":
        return "s-ok"
    if v.startswith("未確認"):
        return "s-unknown"
    return ""


STATUS_COLS_01 = {7, 8, 9, 10, 11, 12, 13}   # 0-based indices in sheet 01 data rows

# ステータス凡例アイコン (✓ 正常 / ▲ 要確認 / ✕ NG / － データなし)
_STATUS_ICON = {"s-ok": "✓", "s-warn": "▲", "s-ng": "✕", "s-unknown": "－"}

# 01_一次承認 の列グループ (見出し2段表示用). 各要素は (グループ名, 列スパン数).
# header 全列を左から連続でカバーする (ラベル空文字は無地スパン).
SHEET01_GROUPS = [
    ("", 4),                          # 伝票No./入力者名/社員番号/所属
    ("出張情報", 2),                   # 出張期間/合計金額
    ("承認状況", 1),                   # 承認状態
    ("1. 出張実態の確認", 1),           # 出張実態
    ("2. 労務・健康管理の確認", 1),      # 労務
    ("3. 出張費・宿泊費上限確認", 1),    # 金額規程
    ("4. 全体チェック", 4),             # 二重申請/領収書/承認ルート/総合判定
    ("詳細", 2),                       # 要確認項目/差戻し候補
]


def _cell_html(val: str, col_idx: int, sheet_id: str, i18n: dict | None = None) -> str:
    cls = ""
    if sheet_id == "01" and col_idx in STATUS_COLS_01:
        cls = _status_class(val)
    elif sheet_id == "03" and col_idx == 3:
        cls = _status_class(val)
    elif sheet_id == "04" and col_idx == 4:
        cls = _status_class(val)
    elif sheet_id == "05" and col_idx == 4:
        cls = "s-ok" if val == "OK" else "s-ng"
    display = _translate_cell_val(val, i18n) if i18n else val
    if cls:
        icon = _STATUS_ICON.get(cls, "")
        return f'<span class="badge {cls}">{icon} {display}</span>'
    return display


def _group_header_html(groups: list, header_len: int, tr: dict) -> str:
    total = sum(span for _, span in groups)
    if total != header_len:
        return ""  # 列数不一致なら安全にグループ行を省略
    cells = []
    start = 0
    for label, span in groups:
        if label:
            cells.append(
                f'<th colspan="{span}" class="grp-th" data-start="{start}" '
                f'data-span="{span}" onclick="highlightGroup(this)" '
                f'onmouseenter="previewGroup(this,true)" onmouseleave="previewGroup(this,false)">'
                f'{tr.get(label, label)}</th>'
            )
        else:
            cells.append(f'<th colspan="{span}" class="grp-th grp-th-blank"></th>')
        start += span
    return f"<tr class='grp-row'>{''.join(cells)}</tr>"


def _build_table(header: list[str], rows: list[list[str]], sheet_id: str,
                 col_translations: dict | None = None, i18n: dict | None = None,
                 groups: list | None = None) -> str:
    tr = col_translations or {}
    group_row_html = _group_header_html(groups, len(header), tr) if groups else ""
    th_html = "".join(
        f'<th onclick="sortTable(this)" data-col="{i}">'
        f'{tr.get(h, h)}<span class="sort-icon">⇅</span></th>'
        for i, h in enumerate(header)
    )
    tr_html_parts = []
    for row in rows:
        tds = "".join(
            f'<td title="{str(row[i]) if i < len(row) else ""}">'
            f'{_cell_html(str(row[i]) if i < len(row) else "", i, sheet_id, i18n)}</td>'
            for i in range(len(header))
        )
        status_attr = ""
        if sheet_id == "01" and len(row) > 13:
            raw_status = str(row[13])
            code = "未確認" if raw_status.startswith("未確認") else raw_status
            status_attr = f' data-status="{code}"'
        tr_html_parts.append(f"<tr{status_attr}>{tds}</tr>")
    tbody = "\n".join(tr_html_parts)
    return f"""
<div class="tbl-wrap">
  <table class="data-tbl" id="tbl-{sheet_id}">
    <thead>{group_row_html}<tr>{th_html}</tr></thead>
    <tbody>{tbody}</tbody>
  </table>
</div>"""


I18N = {
    "ja": {
        "title":        "出張精算 承認チェックシート ビューワー",
        "generated":    "生成日時",
        "search":       "🔍 検索...",
        "count_suffix": "件",
        "lbl_ng":       "NG",
        "lbl_warn":     "要確認",
        "lbl_ok":       "OK",
        "lbl_unknown":  "未確認",
        "lbl_total":    "合計件数",
        "legend_title": "ステータス凡例：",
        "filter_status_all": "ステータス：すべて",
        "tabs": {
            "01": "01_一次承認",
            "02": "02_二次明細",
            "03": "03_差異一覧",
            "04": "04_差戻し文面",
            "05": "05_取込ログ",
            "06": "06_判定ルール",
            "07": "07_マスタ確認",
        },
        "lang_attr": "ja",
        "font": '"Hiragino Kaku Gothic ProN","Yu Gothic",Meiryo,"Noto Sans JP",sans-serif',
    },
    "uz": {
        "title":        "Xizmat safari xarajatlari tasdiqlash varag'i",
        "generated":    "Yaratilgan sana",
        "search":       "🔍 Qidirish...",
        "count_suffix": "ta",
        "lbl_ng":       "NG (Rad)",
        "lbl_warn":     "Tekshirish kerak",
        "lbl_ok":       "OK",
        "lbl_unknown":  "Noaniq",
        "lbl_total":    "Jami ariza",
        "legend_title": "Status belgilari:",
        "filter_status_all": "Status: barchasi",
        "tabs": {
            "01": "01 — Birlamchi tasdiqlash",
            "02": "02 — Tafsilot",
            "03": "03 — Farqlar ro'yxati",
            "04": "04 — Qaytarish matni",
            "05": "05 — Yuklash jurnali",
            "06": "06 — Qoidalar",
            "07": "07 — Master tekshiruvi",
        },
        "lang_attr": "uz",
        "font": '"Inter","Segoe UI",system-ui,sans-serif',
        "prefix_map": [
            ("取込日時 ", "Yuklash vaqti "),
        ],
        "value_map": {
            # ── 承認状態 ──
            "承認済":                         "Tasdiqlangan",
            "承認中":                         "Tasdiqlash jarayonida",
            "未承認":                         "Tasdiqlanmagan",
            "未承認(PENDING)":                "Tasdiqlanmagan",
            "未確認(勤怠データ欠落)":         "Noaniq (davomati yo'q)",
            # ── 判定 ──
            "要確認":                         "Tekshirish kerak",
            "要確認(勤怠データ欠落)":         "Tekshirish kerak (davomati yo'q)",
            # ── Sheet 02 証票 ──
            "証票あり":                       "Kvitansiya bor",
            "証票なし":                       "Kvitansiya yo'q",
            "免除":                           "Mustasnо",
            # ── Sheet 02 照合状態 ──
            "突合":                           "Mos keldi",
            "別名突合":                       "Taxminiy mos",
            "未突合":                         "Mos kelmadi",
            # ── Sheet 03 観点 (compound parts) ──
            "出張実態":                       "Safari holati",
            "労務":                           "Mehnat",
            "金額規程":                       "Summa qoidasi",
            "二重申請":                       "Ikki marta ariza",
            "領収書":                         "Kvitansiya",
            "承認ルート":                     "Tasdiqlash zanjiri",
            # ── 判定理由 (static known strings) ──
            "旅費規定の上限を超過した明細の妥当性を確認してください。":
                "Safar xarajatlari qoidasidan oshgan qatorlarni tekshiring",
            "金額がマイナス/小計不一致の明細を是正してください。":
                "Manfiy summa yoki yig'indi mos kelmaydigan qatorlarni to'g'rilang",
            "深夜帯の移動について時刻・深夜手当の要否をご確認ください。":
                "Kech tun harakati vaqti va tungi to'lov zaruratini tekshiring",
            "承認者ルート判定不可: 申請者がマスタ/名簿に未登録":
                "Tasdiqlash zanjiri aniqlanmadi: ariza beruvchi ro'yxatda yo'q",
            "免除交通機関だが高額の明細に領収書がありません。領収書の添付を依頼してください。":
                "Mustasnо transport bo'lsa-da yuqori summali qatorda kvitansiya yo'q. Kvitansiya so'rang",
            # ── Transport turlari ──
            "電車･ﾊﾞｽ":                      "Poyezd/Avtobus",
            "飛行機":                         "Samolyot",
            "車(同乗)":                       "Mashina (birgalikda)",
            "ﾀｸｼｰ(同乗)":                   "Taksi (birgalikda)",
            "レンタカー":                     "Ijaraga mashina",
            "テレワーク":                     "Masofaviy ish",
            "徒歩":                           "Piyoda",
            "作業･打合せ":                    "Ish/Yig'ilish",
            "ホテル":                         "Mehmonxona",
            # ── Xarajat kategoriyalari ──
            "ホテル代":                       "Mehmonxona to'lovi",
            "宿泊税":                         "Mehmonxona solig'i",
            "入湯税":                         "Hammom solig'i",
            "ガソリン代":                     "Benzin",
            "ｶﾞｿﾘﾝ代":                       "Benzin",
            "コインパーキング":               "Pulli avtoturargoh",
            "駐車代":                         "Avtoturargoh to'lovi",
            "駐車代金":                       "Avtoturargoh to'lovi",
            "駐車場":                         "Avtoturargoh",
            "駐車場代":                       "Avtoturargoh to'lovi",
            "旅費交通費":                     "Sayohat/transport xarajatlari",
            "委託サービス費":                 "Shartnoma xizmat haqi",
            "賃借料":                         "Ijara to'lovi",
            # ── Tekshirish tizimlari ──
            "楽々勤怠 (勤務実績)":            "Davomati tizimi (ish natijalari)",
            "楽々勤怠 (勤務時刻/休日)":       "Davomati tizimi (vaqt/dam olish)",
            "楽々精算 (他申請)":              "Xarajat tizimi (boshqa ariza)",
            "楽々精算 (添付/証票)":           "Xarajat tizimi (ilova/kvitansiya)",
            "楽々精算 + 旅費規定":            "Xarajat tizimi + safari qoidalari",
            "20期承認者名簿 + 楽々精算":      "20-davr tasdiqlovchilar ro'yxati + xarajat tizimi",
            # ── Boshqa ──
            "(集計)":                         "(Yig'indi)",
            "複数候補":                       "Ko'p nomzod",
            # ── Sheet 05 区分 ──
            "出張精算CSV":                    "Xizmat safari CSV",
            "勤怠(出勤簿)":                   "Davomati (ish vaqti kitobi)",
            "社員マスタ":                     "Xodimlar ro'yxati",
            "顧客マスタ":                     "Mijozlar ro'yxati",
            "承認者名簿":                     "Tasdiqlovchilar ro'yxati",
            "件数照合":                       "Sonlar tekshiruvi",
            # ── Sheet 05 結果 ──
            "エラー":                         "Xato",
            # ── Sheet 06 項目 ──
            "氏名ファジー閾値":               "Ism taxminiy chegarasi",
            "地名照合閾値":                   "Joylashuv moslashtirish chegarasi",
            "深夜発 閾値":                    "Kech tun jo'nab ketish chegarasi",
            "深夜着 閾値":                    "Kech tun yetib kelish chegarasi",
            "領収書 高額暫定閾値":            "Kvitansiya yuqori qiymat chegarasi",
            "領収書 検知下限":                "Kvitansiya minimal tekshiruv summasi",
            "金額規程(上限)提供":             "Summa qoidasi (yuqori limit) mavjud",
            "領収書要否規程提供":             "Kvitansiya talab qoidasi mavjud",
            "確認のみ=承認 扱い":             "Faqat 'tekshirish' = tasdiqlash sanaladi",
            "出張日当_一般職":                "Safari kunlik to'lov — oddiy xodim",
            "出張日当_管理職":                "Safari kunlik to'lov — boshqaruv",
            "滞在補助費_一般職":              "Qolish yordami — oddiy xodim",
            "滞在補助費_管理職":              "Qolish yordami — boshqaruv",
            "出張加算日当_一般職":            "Safari qo'shimcha kunlik — oddiy xodim",
            "出張加算日当_主任以上":          "Safari qo'shimcha kunlik — katta mutaxassis+",
            "ホテル代_東京23区_管理職":       "Mehmonxona (Tokio 23 tuman) — boshqaruv",
            "ホテル代_東京23区_一般職":       "Mehmonxona (Tokio 23 tuman) — oddiy xodim",
            "ホテル代_その他_管理職":         "Mehmonxona (boshqa hududlar) — boshqaruv",
            "ホテル代_その他_一般職":         "Mehmonxona (boshqa hududlar) — oddiy xodim",
            "既知の前提/欠落":                "Ma'lum ogohlantirishlar",
            # ── Sheet 06 値 ──
            "あり":                           "mavjud",
            "しない":                         "yo'q",
            # ── Sheet 06 備考 ──
            "difflib ratio (社員突合)":        "difflib ratio (xodim moslashtirish)",
            "rapidfuzz partial_ratio (顧客突合)": "rapidfuzz partial_ratio (mijoz moslashtirish)",
            "移動開始がこれ以前":              "Harakat boshlash vaqti bundan oldin bo'lsa",
            "移動終了がこれ以降":              "Harakat tugash vaqti bundan keyin bo'lsa",
            "規程未提供時の高額判定":          "Qoida taqdim etilmaganda yuqori summa tekshiruvi",
            "これ未満の小額は要確認にしない":  "Bundan kam kichik summalar tekshirilmaydi",
            "未提供時は金額上限を要確認(NGは出さない)":
                "Taqdim etilmaganda summa limiti tekshiruv (NG chiqarilmaydi)",
            "未提供時は暫定閾値で判定":        "Taqdim etilmaganda vaqtinchalik chegara ishlatiladi",
            "(確認)承認者を正式承認と見なすか":
                "(Tekshirish) tasdiqlovchi rasmiy tasdiqlash sanalsinmi",
            "円/日":                          "¥/kun",
            "円":                             "¥",
            # known_gaps strings (sheet 06 備考 column)
            "2026-05 の出勤簿(勤怠)が未提供 — 出張実態・労務の勤怠照合は劣化(advisory)。":
                "2026-05 oy davomati taqdim etilmagan — safari holati va mehnat tekshiruvi sifati pasaygan (advisory)",
            "社員マスタに役職列(役職/グレード/職位)が未登録の場合、役職不明として一般職上限と管理職上限の間は要確認扱いになります。":
                "Xodim ro'yxatida lavozim ustuni bo'lmasa, lavozim noaniq sifatida oddiy xodim va boshqaruv limiti orasidagi summa tekshiruv talab qiladi",
            # ── Sheet 07 種別 ──
            "承認者名簿 未登録":              "Tasdiqlovchilar ro'yxatida yo'q",
            "名簿/マスタ不整合":              "Ro'yxat/Master nomuvofiqlik",
            # ── Sheet 07 詳細 ──
            "申請者が20期承認者名簿に未登録 (出張命令者を判定不可)":
                "Ariza beruvchi 20-davr tasdiqlovchilar ro'yxatida yo'q (safari buyrug'i beruvchini aniqlab bo'lmaydi)",
            "20期名簿に存在するが社員マスタに無い氏名":
                "20-davr ro'yxatida bor, lekin xodimlar ro'yxatida yo'q ism",
            # ── Sheet 07 対応 ──
            "20期名簿に申請者を登録":
                "20-davr ro'yxatiga ariza beruvchini qo'shing",
            "社員マスタ/名簿の氏名表記を突合・統一":
                "Xodimlar ro'yxati/nomlar ro'yxatidagi ism yozuvini solishtiring va birlashtiring",
            # ── 対応案 / 差戻し文面 (static suggestion strings) ──
            "承認手続きが未完了です。承認実行のうえ再提出してください。":
                "Tasdiqlash jarayoni tugallanmagan. Tasdiqlang va qayta yuboring.",
            "申請者を従業員マスタ/承認者名簿(20期)に登録のうえ再判定してください。":
                "Ariza beruvchini xodimlar/tasdiqlovchilar ro'yxatiga (20-davr) qo'shing va qayta tekshiring.",
            "非労働日/休暇日の移動について出張の必要性・実態をご確認ください。":
                "Dam olish/ta'til kunidagi harakat uchun safari zaruratini tekshiring.",
            "該当月の勤怠データが未提供のため労務照合は劣化しています。勤怠提供後に再確認してください。":
                "Tegishli oy davomati taqdim etilmagan — mehnat tekshiruvi sifati pasaygan. Davomati kelib tushgach qayta tekshiring.",
            "労務 (勤怠整合) について手動でご確認ください。":
                "Mehnat (davomat muvofiqligi) ni qo'lda tekshiring.",
            "同一日・同一区間・同額の明細が他伝票と重複していないか確認し、必要なら一方を取消。":
                "Bir xil sana/marshrut/summa bilan boshqa hujjatda takror qator bor-yo'qligini tekshiring; kerak bo'lsa birini bekor qiling.",
            "同日・同一顧客先への複数申請の妥当性を確認。":
                "Bir kunda bir xil mijozga bir nechta ariza asosliligini tekshiring.",
            "明細金額の合計が申告合計を超過しています。金額を確認・修正してください。":
                "Qatorlar summasi e'lon qilingan umumiy summadan oshib ketgan. Miqdorni tekshiring.",
            "申告合計と明細合計の差額(手当コードなし)の内訳を確認してください。":
                "E'lon qilingan va qatorlar summasi farqi (to'lovsiz) sababini tekshiring.",
            "行数(宣言)と明細件数の差異を確認してください。":
                "E'lon qilingan qatorlar soni bilan haqiqiy qatorlar soni mos kelmaydi.",
            "金額がマイナス/小計不一致の明細を是正してください。":
                "Manfiy summa yoki yig'indi mos kelmaydigan qatorlarni to'g'rilang.",
            "旅費規定の上限を超過した明細の妥当性を確認してください。":
                "Safar xarajatlari qoidasidan oshgan qatorlarni tekshiring.",
        },
        "notice_map": {
            "2026-05 の出勤簿(勤怠)が未提供":
                "・2026-05 oy davomati (ish vaqti) taqdim etilmagan — safari holati va mehnat tekshiruvi sifati pasaygan (advisory)",
            "社員マスタに役職列":
                "・Xodim ro'yxatida lavozim ustuni (lavozim/daraja/mansab) ro'yxatga olinmagan bo'lsa, lavozim noaniq sifatida oddiy xodim va boshqaruv limiti orasidagi summa tekshiruv talab qiladi",
            "勤怠データ対象月":
                "・Davomati hisobot oylari",
        },
        "headers": {
            # ── Sheet 01 列グループ見出し ──
            "出張情報":                       "Safar ma'lumoti",
            "承認状況":                       "Tasdiqlash holati",
            "1. 出張実態の確認":              "1. Safar holati tekshiruvi",
            "2. 労務・健康管理の確認":        "2. Mehnat/sog'liq tekshiruvi",
            "3. 出張費・宿泊費上限確認":      "3. Xarajat/mehmonxona limiti",
            "4. 全体チェック":                "4. Umumiy tekshiruv",
            "詳細":                           "Tafsilot",
            # ── Sheet 01 ──
            "伝票No.":      "Hujjat №",
            "入力者名":     "Murojaat qiluvchi",
            "社員番号":     "Xodim raqami",
            "所属":         "Bo'lim",
            "出張期間":     "Safari muddati",
            "合計金額":     "Jami summa",
            "承認状態":     "Tasdiqlash holati",
            "出張実態":     "Safari holati",
            "労務":         "Mehnat",
            "金額規程":     "Summa qoidasi",
            "二重申請":     "Ikki marta ariza",
            "領収書":       "Kvitansiya",
            "承認ルート":   "Tasdiqlash zanjiri",
            "総合判定":     "Umumiy baholash",
            "要確認項目":   "Tekshirilishi kerak",
            "差戻し候補":   "Qaytarish sababi",
            # ── Sheet 02 ──
            "明細No.":      "Qator №",
            "明細日付":     "Sana",
            "開始":         "Boshlanish",
            "終了":         "Tugash",
            "出発地":       "Jo'nab ketish joyi",
            "到着地":       "Borish joyi",
            "交通機関":     "Transport",
            "金額":         "Summa",
            "証票":         "Kvitansiya holati",
            "日当CD":       "Kunlik to'lov kodi",
            "宿泊CD":       "Yotoqxona kodi",
            "滞在CD":       "Qolish kodi",
            "勘定科目名":   "Hisobvaraq nomi",
            "照合顧客名":   "Moslashtirish mijozi",
            "距離区分":     "Masofa turi",
            "照合状態":     "Moslashtirish holati",
            "複数候補":     "Ko'p nomzod",
            # ── Sheet 03 ──
            "観点":         "Tekshiruv nuqtai nazari",
            "判定":         "Baholash",
            "判定理由":     "Baholash sababi",
            "確認先システム": "Tekshirish tizimi",
            "対応案":       "Tavsiya",
            # ── Sheet 04 ──
            "宛先(メール)": "Qabul qiluvchi (email)",
            "理由区分":     "Sabab turi",
            "差戻し文面候補": "Qaytarish matni",
            # ── Sheet 05 ──
            "区分":         "Tur",
            "ファイル名":   "Fayl nomi",
            "件数":         "Soni",
            "詳細":         "Tafsilot",
            "結果":         "Natija",
            # ── Sheet 06 ──
            "項目":         "Parametr",
            "値":           "Qiymat",
            "備考":         "Izoh",
            # ── Sheet 07 ──
            "種別":         "Tur",
            "対象":         "Ob'ekt",
            "対応":         "Chora",
        },
    },
}


def _sheet01_stats(rows: list[list[str]], i18n: dict) -> str:
    ng      = sum(1 for r in rows if len(r) > 13 and r[13] == "NG")
    warn    = sum(1 for r in rows if len(r) > 13 and r[13] == "要確認")
    ok      = sum(1 for r in rows if len(r) > 13 and r[13] == "OK")
    unknown = sum(1 for r in rows if len(r) > 13 and r[13].startswith("未確認"))
    legend = f"""
<div class="legend-row">
  <span class="legend-title">{i18n["legend_title"]}</span>
  <span class="legend-item"><span class="badge s-ok">✓ {i18n["lbl_ok"]}</span></span>
  <span class="legend-item"><span class="badge s-warn">▲ {i18n["lbl_warn"]}</span></span>
  <span class="legend-item"><span class="badge s-ng">✕ {i18n["lbl_ng"]}</span></span>
  <span class="legend-item"><span class="badge s-unknown">－ {i18n["lbl_unknown"]}</span></span>
</div>"""
    return legend + f"""
<div class="stats-row">
  <div class="stat-box s-ng-box"><div class="stat-num">{ng}</div><div class="stat-lbl">{i18n["lbl_ng"]}</div></div>
  <div class="stat-box s-warn-box"><div class="stat-num">{warn}</div><div class="stat-lbl">{i18n["lbl_warn"]}</div></div>
  <div class="stat-box s-ok-box"><div class="stat-num">{ok}</div><div class="stat-lbl">{i18n["lbl_ok"]}</div></div>
  <div class="stat-box s-unk-box"><div class="stat-num">{unknown}</div><div class="stat-lbl">{i18n["lbl_unknown"]}</div></div>
  <div class="stat-box s-total-box"><div class="stat-num">{len(rows)}</div><div class="stat-lbl">{i18n["lbl_total"]}</div></div>
</div>"""


def _render_sheet(sid: str, label: str, data: dict, i18n: dict) -> str:
    header = data["header"]
    rows = data["rows"]
    notices_html = ""
    if "notices" in data and data["notices"]:
        notice_map = i18n.get("notice_map", {})
        def _translate_notice(n: str) -> str:
            for ja_key, uz_val in notice_map.items():
                if ja_key in n:
                    if "対象月" in ja_key:
                        suffix = n.split(":")[-1].strip() if ":" in n else ""
                        return uz_val + (f": {suffix}" if suffix else "")
                    return uz_val
            return n
        items = "".join(f"<li>{_translate_notice(n)}</li>" for n in data["notices"])
        notices_html = f'<div class="notice-box"><ul>{items}</ul></div>'
    stats_html = _sheet01_stats(rows, i18n) if sid == "01" else ""
    suffix = i18n["count_suffix"]
    if sid == "01":
        status_opts = "".join(
            f'<option value="{v}">{lbl}</option>'
            for v, lbl in [
                ("OK", i18n["lbl_ok"]), ("要確認", i18n["lbl_warn"]),
                ("NG", i18n["lbl_ng"]), ("未確認", i18n["lbl_unknown"]),
            ]
        )
        search_html = f"""
<div class="toolbar">
  <input class="search-box" type="text" placeholder="{i18n['search']}" oninput="applyFilters01('{suffix}')" id="search-{sid}">
  <select class="status-select" id="status-{sid}" onchange="applyFilters01('{suffix}')">
    <option value="">{i18n['filter_status_all']}</option>
    {status_opts}
  </select>
  <span class="row-count" id="count-{sid}">{len(rows)} {suffix}</span>
</div>"""
    else:
        search_html = f"""
<div class="toolbar">
  <input class="search-box" type="text" placeholder="{i18n['search']}" oninput="filterTable(this, '{sid}', '{suffix}')">
  <span class="row-count" id="count-{sid}">{len(rows)} {suffix}</span>
</div>"""
    groups = SHEET01_GROUPS if sid == "01" else None
    table_html = _build_table(header, rows, sid, i18n.get("headers"), i18n, groups=groups)
    return f"""
<div id="panel-{sid}" class="panel" style="display:none">
  {notices_html}
  {stats_html}
  {search_html}
  {table_html}
</div>"""


CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
:root {
  --navy: #1b3a6b;
  --navy2: #14305a;
  --accent: #2563eb;
  --ok: #166534;
  --ok-bg: #dcfce7;
  --ng: #991b1b;
  --ng-bg: #fee2e2;
  --warn: #92400e;
  --warn-bg: #fef3c7;
  --unk: #374151;
  --unk-bg: #f3f4f6;
  --border: #d1d5db;
  --stripe: #f8fafc;
  --text: #111827;
  --radius: 6px;
}
body {
  font-family: "Hiragino Kaku Gothic ProN","Yu Gothic",Meiryo,"Noto Sans JP",sans-serif;
  font-size: 12.5px;
  color: var(--text);
  background: #f1f5f9;
  min-height: 100vh;
}
/* ── Header ── */
.app-header {
  background: var(--navy);
  color: #fff;
  padding: 14px 24px 0;
  position: sticky;
  top: 0;
  z-index: 100;
  box-shadow: 0 2px 8px rgba(0,0,0,0.25);
}
.app-title { font-size: 15px; font-weight: 700; letter-spacing: 0.04em; margin-bottom: 10px; }
.app-meta { font-size: 11px; color: rgba(255,255,255,0.65); margin-bottom: 10px; }
/* ── Tabs ── */
.tabs { display: flex; gap: 2px; }
.tab-btn {
  padding: 8px 14px;
  background: rgba(255,255,255,0.12);
  color: rgba(255,255,255,0.75);
  border: none;
  border-radius: 6px 6px 0 0;
  cursor: pointer;
  font-size: 12px;
  font-family: inherit;
  white-space: nowrap;
  transition: background 0.15s;
}
.tab-btn:hover { background: rgba(255,255,255,0.22); color:#fff; }
.tab-btn.active { background: #fff; color: var(--navy); font-weight: 700; }
/* ── Main content ── */
.main { padding: 20px 24px; }
.panel { animation: fadeIn 0.15s ease; }
@keyframes fadeIn { from { opacity:0; transform:translateY(4px); } to { opacity:1; transform:none; } }
/* ── Notice box ── */
.notice-box {
  background: #fff8e1;
  border-left: 3px solid #f59e0b;
  border-radius: 0 var(--radius) var(--radius) 0;
  padding: 10px 14px;
  margin-bottom: 14px;
  font-size: 12px;
  color: #78350f;
}
.notice-box ul { padding-left: 16px; }
.notice-box li { margin-bottom: 3px; }
/* ── Legend ── */
.legend-row {
  display: flex; align-items: center; gap: 14px; flex-wrap: wrap;
  margin-bottom: 10px; font-size: 12px; color: #4b5563;
}
.legend-title { font-weight: 600; }
.legend-item { display: inline-flex; align-items: center; }
/* ── Stats ── */
.stats-row { display: flex; gap: 10px; margin-bottom: 16px; flex-wrap: wrap; }
.stat-box {
  flex: 1; min-width: 100px;
  border-radius: var(--radius);
  padding: 12px 16px;
  text-align: center;
  border: 1px solid var(--border);
}
.stat-num { font-size: 28px; font-weight: 700; line-height: 1; }
.stat-lbl { font-size: 11px; margin-top: 4px; }
.s-ng-box   { background: var(--ng-bg);  color: var(--ng);   border-color: #fca5a5; }
.s-warn-box { background: var(--warn-bg); color: var(--warn); border-color: #fcd34d; }
.s-ok-box   { background: var(--ok-bg);  color: var(--ok);   border-color: #86efac; }
.s-unk-box  { background: var(--unk-bg); color: var(--unk);  border-color: #d1d5db; }
.s-total-box{ background: #eff6ff; color: var(--accent); border-color: #bfdbfe; }
/* ── Toolbar ── */
.toolbar { display: flex; align-items: center; gap: 12px; margin-bottom: 10px; }
.search-box {
  padding: 7px 12px;
  border: 1px solid var(--border);
  border-radius: var(--radius);
  font-size: 12.5px;
  font-family: inherit;
  width: 320px;
  background: #fff;
}
.search-box:focus { outline: none; border-color: var(--accent); box-shadow: 0 0 0 3px rgba(37,99,235,0.15); }
.status-select {
  padding: 7px 10px;
  border: 1px solid var(--border);
  border-radius: var(--radius);
  font-size: 12.5px;
  font-family: inherit;
  background: #fff;
}
.status-select:focus { outline: none; border-color: var(--accent); box-shadow: 0 0 0 3px rgba(37,99,235,0.15); }
.row-count { font-size: 11px; color: #6b7280; }
/* ── Table ── */
.tbl-wrap {
  overflow-x: auto;
  border-radius: var(--radius);
  box-shadow: 0 1px 4px rgba(0,0,0,0.1);
  background: #fff;
}
.data-tbl { width: 100%; border-collapse: collapse; font-size: 12px; }
.data-tbl thead tr { background: var(--navy); color: #fff; }
.data-tbl thead th {
  padding: 9px 12px;
  text-align: left;
  white-space: nowrap;
  cursor: pointer;
  user-select: none;
  font-weight: 600;
  letter-spacing: 0.02em;
}
.data-tbl thead th:hover { background: var(--navy2); }
.sort-icon { margin-left: 4px; opacity: 0.5; font-size: 10px; }
/* ── Grouped header (sheet 01) ── */
.data-tbl thead tr.grp-row { background: #0e2144; }
.data-tbl thead th.grp-th {
  height: 38px;
  padding: 10px 16px;
  text-align: center;
  font-size: 10.5px;
  font-weight: 700;
  letter-spacing: 0.03em;
  border-left: 1px solid rgba(255,255,255,0.2);
  border-bottom: 2px solid rgba(255,255,255,0.35);
  cursor: pointer;
  white-space: nowrap;
  transition: background 0.12s;
}
.data-tbl thead th.grp-th:hover { background: #1e4a8a; }
.data-tbl thead th.grp-th.grp-active { background: var(--accent); }
.data-tbl thead th.grp-th-blank { background: transparent; border-left: none; border-bottom: none; cursor: default; }
.data-tbl thead th.grp-th-blank:hover { background: transparent; }
.data-tbl tbody td.col-hl { background: #dbeafe; }
.data-tbl tbody tr:hover td.col-hl { background: #bfdbfe; }
.data-tbl tbody td.col-hl-preview { background: #eff6ff; }
.data-tbl tbody tr:hover td.col-hl-preview { background: #dbeafe; }
.data-tbl tbody tr:nth-child(even) { background: var(--stripe); }
.data-tbl tbody tr:hover { background: #eff6ff; }
.data-tbl tbody td {
  padding: 6px 12px;
  border-bottom: 1px solid #e5e7eb;
  vertical-align: middle;
  white-space: nowrap;
  overflow: hidden;
  text-overflow: ellipsis;
  max-width: 220px;
}
.data-tbl tbody tr.hidden { display: none; }
/* ── Badges ── */
.badge { display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; white-space: nowrap; }
.s-ng      { background: var(--ng-bg);   color: var(--ng);   }
.s-warn    { background: var(--warn-bg); color: var(--warn); }
.s-ok      { background: var(--ok-bg);   color: var(--ok);   }
.s-unknown { background: var(--unk-bg);  color: var(--unk);  }
"""

JS = """
// ── Tab switching ──
function showTab(id) {
  document.querySelectorAll('.panel').forEach(p => p.style.display = 'none');
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  document.getElementById('panel-' + id).style.display = 'block';
  document.querySelector('[data-tab="' + id + '"]').classList.add('active');
}

// ── Search / filter ──
function filterTable(input, sid, suffix) {
  suffix = suffix || '件';
  const q = input.value.toLowerCase();
  const tbody = document.querySelector('#tbl-' + sid + ' tbody');
  let visible = 0;
  tbody.querySelectorAll('tr').forEach(tr => {
    const match = tr.textContent.toLowerCase().includes(q);
    tr.classList.toggle('hidden', !match);
    if (match) visible++;
  });
  document.getElementById('count-' + sid).textContent = visible + ' ' + suffix;
}

// ── Sheet 01: text search + status dropdown combined ──
function applyFilters01(suffix) {
  suffix = suffix || '件';
  const q = document.getElementById('search-01').value.toLowerCase();
  const status = document.getElementById('status-01').value;
  const tbody = document.querySelector('#tbl-01 tbody');
  let visible = 0;
  tbody.querySelectorAll('tr').forEach(tr => {
    const textMatch = tr.textContent.toLowerCase().includes(q);
    const statusMatch = !status || tr.dataset.status === status;
    const match = textMatch && statusMatch;
    tr.classList.toggle('hidden', !match);
    if (match) visible++;
  });
  document.getElementById('count-01').textContent = visible + ' ' + suffix;
}

// ── Group header hover: temporary column preview (removed on mouseleave) ──
function previewGroup(th, on) {
  const table = th.closest('table');
  const start = parseInt(th.dataset.start);
  const span = parseInt(th.dataset.span);
  table.querySelectorAll('tbody tr').forEach(tr => {
    for (let i = start; i < start + span; i++) {
      if (tr.cells[i]) tr.cells[i].classList.toggle('col-hl-preview', on);
    }
  });
}

// ── Group header click: highlight its columns ──
function highlightGroup(th) {
  const table = th.closest('table');
  const wasActive = th.classList.contains('grp-active');
  table.querySelectorAll('.grp-th').forEach(t => t.classList.remove('grp-active'));
  table.querySelectorAll('td.col-hl').forEach(td => td.classList.remove('col-hl'));
  if (wasActive) return;
  th.classList.add('grp-active');
  const start = parseInt(th.dataset.start);
  const span = parseInt(th.dataset.span);
  table.querySelectorAll('tbody tr').forEach(tr => {
    for (let i = start; i < start + span; i++) {
      if (tr.cells[i]) tr.cells[i].classList.add('col-hl');
    }
  });
}

// ── Sort ──
let _sortState = {};
function sortTable(th) {
  const table = th.closest('table');
  const sid = table.id.replace('tbl-', '');
  const col = parseInt(th.dataset.col);
  const key = sid + '-' + col;
  const asc = !_sortState[key];
  _sortState[key] = asc;

  // reset icons
  th.closest('thead').querySelectorAll('.sort-icon').forEach(s => s.textContent = '⇅');
  th.querySelector('.sort-icon').textContent = asc ? '↑' : '↓';

  const tbody = table.querySelector('tbody');
  const rows = Array.from(tbody.querySelectorAll('tr'));
  rows.sort((a, b) => {
    const av = a.cells[col] ? a.cells[col].textContent.trim() : '';
    const bv = b.cells[col] ? b.cells[col].textContent.trim() : '';
    const an = parseFloat(av.replace(/[^0-9.-]/g, ''));
    const bn = parseFloat(bv.replace(/[^0-9.-]/g, ''));
    if (!isNaN(an) && !isNaN(bn)) return asc ? an - bn : bn - an;
    return asc ? av.localeCompare(bv, 'ja') : bv.localeCompare(av, 'ja');
  });
  rows.forEach(r => tbody.appendChild(r));
}

// init
showTab('01');
"""


def write_html(sheets: dict, out_path: str, lang: str = "ja") -> None:
    """sheets: {'01': {header, rows, notices?}, '02': ..., ...}"""
    i18n = I18N.get(lang, I18N["ja"])
    tab_labels = i18n["tabs"]

    tabs_html = "".join(
        f'<button class="tab-btn" data-tab="{sid}" onclick="showTab(\'{sid}\')">{label}</button>'
        for sid, label in tab_labels.items()
        if sid in sheets
    )

    panels_html = "".join(
        _render_sheet(sid, tab_labels.get(sid, sid), sheets[sid], i18n)
        for sid in tab_labels
        if sid in sheets
    )

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    font  = i18n["font"]
    css   = CSS.replace(
        '"Hiragino Kaku Gothic ProN","Yu Gothic",Meiryo,"Noto Sans JP",sans-serif',
        font,
    )

    html = f"""<!doctype html>
<html lang="{i18n['lang_attr']}">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>{i18n['title']}</title>
<style>{css}</style>
</head>
<body>
<div class="app-header">
  <div class="app-title">{i18n['title']}</div>
  <div class="app-meta">{i18n['generated']}: {stamp}</div>
  <div class="tabs">{tabs_html}</div>
</div>
<div class="main">
  {panels_html}
</div>
<script>{JS}</script>
</body>
</html>"""

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)


def read_excel_and_write_html(xlsx_path: str, html_path: str, lang: str = "ja") -> None:
    """既存の Excel ファイルを読んで HTML を生成する."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheets: dict = {}

    sheet_map = {
        "01_一次承認チェック": "01",
        "02_二次承認詳細": "02",
        "03_差異一覧": "03",
        "04_差戻し文面候補": "04",
        "05_取込ログ": "05",
        "06_判定ルール": "06",
        "07_マスタ確認": "07",
    }

    for sheet_name, sid in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))

        if sid == "01":
            notices = []
            header_idx = 0
            for i, row in enumerate(all_rows[:6]):
                v = row[0]
                if v and str(v).startswith("・"):
                    notices.append(str(v))
                elif v and "伝票" in str(v):
                    header_idx = i
                    break
            header = [str(v) if v is not None else "" for v in all_rows[header_idx]]
            data_rows = [
                [str(v) if v is not None else "" for v in r]
                for r in all_rows[header_idx + 1:]
                if any(v is not None for v in r)
            ]
            sheets[sid] = {"header": header, "rows": data_rows, "notices": notices}
        else:
            header = [str(v) if v is not None else "" for v in all_rows[0]]
            data_rows = [
                [str(v) if v is not None else "" for v in r]
                for r in all_rows[1:]
                if any(v is not None for v in r)
            ]
            sheets[sid] = {"header": header, "rows": data_rows}

    write_html(sheets, html_path, lang=lang)
    print(f"HTML 出力 [{lang}]: {html_path}")


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python3 html_writer_web.py <excel_path> [html_path]")
        sys.exit(1)
    xlsx = sys.argv[1]
    html = sys.argv[2] if len(sys.argv) > 2 else xlsx.replace(".xlsx", ".html")
    read_excel_and_write_html(xlsx, html)
