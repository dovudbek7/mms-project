"""承認者ロスタ (20期評価者・承認者一覧) ローダ.

データ契約 §(f). '20期' シートから ApproverRule (氏名 -> 出張命令者 等) を構築する.
特徴:
  - 部署/課/グループ (B/C/D列) は縦方向に結合セルで表現される. read_only を
    使わずに開き merged_cells.ranges を展開して各行の所属を解決する.
  - 出張命令者 (F列) が判定の主対象. '高橋昭太（確認）' のような確認印は
    norm_approver() で (正規化名, 確認のみフラグ) に分解する.
  - 勤怠管理承認者 (G列) は '-'/空欄を None として扱う.
"""
from __future__ import annotations

import re
import unicodedata

import openpyxl

from models import ApproverRule
from normalize import norm, norm_approver

# 部署/課/グループ ラベルから部門コード (NNN) を抽出 (全角/半角括弧に寛容)
_DEPT_CODE_RE = re.compile(r"[（(]\s*(\d+)\s*[）)]")


def _clean_label(v) -> str | None:
    """セル値を表示用ラベルに整形 (NFKC, 改行→空白, 前後空白除去). 空は None."""
    if v is None:
        return None
    s = unicodedata.normalize("NFKC", str(v))
    s = s.replace("\n", " ").replace("\r", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s or None


def _parse_dept_code(*labels) -> str | None:
    """所属ラベル群から最初に見つかった部門コード (NNN) を返す.

    粒度の細かい順 (グループ→課→部署) に走査する想定で呼び出す.
    """
    for lab in labels:
        if not lab:
            continue
        m = _DEPT_CODE_RE.search(unicodedata.normalize("NFKC", str(lab)))
        if m:
            return m.group(1)
    return None


def load_approver_rules(path: str, sheet: str = "20期") -> list[ApproverRule]:
    """20期ロスタを読み, 氏名->出張命令者 等の ApproverRule リストを返す.

    結合セル (部署B/課C/グループD列) を解決するため read_only=False で開く.
    データ行は 6..63 (1-based). E列(氏名)が空の行は区切りとして読み飛ばす.
    期待件数: 53.
    """
    # 結合セル範囲が必要なので read_only は使わない (data_only で計算値を取得)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]

    # --- 結合セルを「起点列ごと」に展開 ---------------------------------
    # B(2)起点の結合 = 部署 (B:D を横断する場合あり), C(3)起点 = 課, D(4)起点 = グループ.
    # 起点列で意味を切り分けることで, B:D 横断結合が C/D を汚染しないようにする.
    dept_fill: dict[int, object] = {}
    sect_fill: dict[int, object] = {}
    team_fill: dict[int, object] = {}
    for rng in ws.merged_cells.ranges:
        top_left = ws.cell(row=rng.min_row, column=rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            if rng.min_col == 2:        # B起点 → 部署
                dept_fill[r] = top_left
            elif rng.min_col == 3:      # C起点 → 課
                sect_fill[r] = top_left
            elif rng.min_col == 4:      # D起点 → グループ
                team_fill[r] = top_left

    def cell_dept(r: int):
        return dept_fill[r] if r in dept_fill else ws.cell(row=r, column=2).value

    def cell_sect(r: int):
        return sect_fill[r] if r in sect_fill else ws.cell(row=r, column=3).value

    def cell_team(r: int):
        return team_fill[r] if r in team_fill else ws.cell(row=r, column=4).value

    rules: list[ApproverRule] = []
    last_dept: str | None = None  # 部署は配下の課/グループ行へ縦持ち (forward-fill)

    for r in range(6, 64):  # iter_rows(min_row=6, max_row=63) 相当
        name_cell = ws.cell(row=r, column=5).value      # E列 = 氏名

        dept = _clean_label(cell_dept(r))
        section = _clean_label(cell_sect(r))
        team = _clean_label(cell_team(r))

        # 新しい部署ラベルが現れたら縦持ちを更新
        if dept is not None:
            last_dept = dept
        else:
            # 部署が空でも配下の課/グループ行なら直近部署を継承
            dept = last_dept if (section or team) else None

        # 氏名が空の行 (9,11,23,28,33 等の区切り) は読み飛ばす
        if name_cell is None or not norm(name_cell):
            continue

        name_raw = str(name_cell)
        name_norm = norm(name_raw)

        # F列 = 出張命令者 (判定の主対象)
        approver_cell = ws.cell(row=r, column=6).value
        trip_raw = "" if approver_cell is None else str(approver_cell)
        trip_norm, trip_confirm = norm_approver(approver_cell)

        # G列 = 勤怠管理承認者 ('-'/空欄 → None)
        att_cell = ws.cell(row=r, column=7).value
        att_str = unicodedata.normalize("NFKC", str(att_cell)).strip() if att_cell is not None else ""
        if att_str in ("", "-"):
            att_raw = None
            att_norm = None
        else:
            att_raw = str(att_cell)
            att_norm, _ = norm_approver(att_cell)

        # 部門コードは細粒度優先 (グループ→課→部署) で抽出
        dept_code = _parse_dept_code(team, section, dept)

        rules.append(
            ApproverRule(
                employee_name_raw=name_raw,
                employee_name_norm=name_norm,
                trip_approver_raw=trip_raw,
                trip_approver_norm=trip_norm,
                trip_approver_is_confirm_only=trip_confirm,
                attendance_approver_raw=att_raw,
                attendance_approver_norm=att_norm,
                department=dept,
                section=section,
                team=team,
                dept_code=dept_code,
            )
        )

    return rules
