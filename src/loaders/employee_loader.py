"""従業員マスタ ローダ (employee master loader).

暗号化 xlsx を msoffcrypto で復号し, シート 'Sheet1' から Employee を構築する.
ヘッダ行 (row 0): 社員番号, 氏名, メールアドレス, 所属部署, 所属管轄.
列はインデックスではなくヘッダ名で選択する.
"""
from __future__ import annotations

import io

import msoffcrypto
import openpyxl

from models import Employee
from normalize import norm


def _decrypt_to_bytesio(path: str, password: str) -> io.BytesIO:
    """暗号化 xlsx をパスワードで復号し BytesIO に展開して返す."""
    buf = io.BytesIO()
    with open(path, "rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password=password)
        office.decrypt(buf)
    buf.seek(0)
    return buf


def load_employees(path: str, password: str) -> list[Employee]:
    """従業員マスタ (暗号化 xlsx) を読み込み Employee のリストを返す.

    手順:
      1. msoffcrypto で復号 -> BytesIO.
      2. openpyxl(read_only=True, data_only=True) で 'Sheet1' を開く.
      3. ヘッダ行で列位置を解決し, データ各行から Employee を構築.
    空行 (社員番号・氏名がともに空) はスキップする.
    """
    buf = _decrypt_to_bytesio(path, password)
    wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
    # 単一シート前提だが将来のタブ名変更に強くするため先頭シートを採る
    # (customer_loader と同じ防御的方針).
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)

    # --- ヘッダ行から列インデックスを解決 (ヘッダ名で選択) ---
    header = next(rows, None)
    if header is None:
        return []
    col = {}
    for i, h in enumerate(header):
        key = norm(h)
        if key:
            col[key] = i

    # 役職列: 役職/グレード/職位 のいずれかを探す (無ければ None)
    _grade_col = next(
        (k for k in col if "役職" in str(k) or "グレード" in str(k) or "職位" in str(k)),
        None,
    )

    def _cell(r, name: str):
        idx = col.get(name)
        if idx is None or idx >= len(r):
            return None
        return r[idx]

    def _text(v) -> str | None:
        """セル値を文字列化して trim. 空は None."""
        if v is None:
            return None
        s = str(v).strip()
        return s or None

    employees: list[Employee] = []
    for r in rows:
        if r is None:
            continue
        emp_id = _text(_cell(r, "社員番号"))
        name_raw = _cell(r, "氏名")
        name_str = _text(name_raw)
        # 社員番号・氏名がともに空の行は空行としてスキップ
        if not emp_id and not name_str:
            continue

        employees.append(
            Employee(
                employee_id=str(emp_id).strip() if emp_id is not None else "",
                name_raw=name_str if name_str is not None else "",
                name_norm=norm(name_raw),
                email=_text(_cell(r, "メールアドレス")),
                department=_text(_cell(r, "所属部署")),  # 空欄は None
                jurisdiction=_text(_cell(r, "所属管轄")),
                grade=_text(_cell(r, _grade_col)) if _grade_col else None,
            )
        )

    return employees
