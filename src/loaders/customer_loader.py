"""顧客マスタローダ (customer master loader).

暗号化 xlsx を msoffcrypto で復号 → openpyxl で読込, Customer レコード列を返す.
データ契約 §1(e) の Customer 型に対応.

ヘッダ (行0): 取引先名, 取引先番号, 距離区分, 都道府県, 東西.
取引先名 は normalize_place() で (ベース, 拠点トークン) に分解する.
距離区分 は parse_band() で (下限, 上限) に分解する ('500~ km' は上限 None).
"""
from __future__ import annotations

import io

import msoffcrypto
import openpyxl

from models import Customer
from normalize import normalize_place, parse_band


def _find_col(header: tuple, token: str) -> int | None:
    """ヘッダ行から token を含む列の index を返す (部分一致, 防御的)."""
    for i, h in enumerate(header):
        if h is not None and token in str(h):
            return i
    return None


def load_customers(path: str, password: str) -> list[Customer]:
    """顧客マスタを復号・読込して Customer のリストを返す.

    path: 暗号化 xlsx パス, password: 復号パスワード.
    最初のシート, ヘッダ行0. 空行 (取引先名・取引先番号ともに空) はスキップ.
    """
    # --- 復号 (msoffcrypto → BytesIO) ---
    decrypted = io.BytesIO()
    with open(path, "rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password=password)
        office.decrypt(decrypted)
    decrypted.seek(0)

    # --- openpyxl 読込 (最初のシート) ---
    wb = openpyxl.load_workbook(decrypted, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return []

    # ヘッダ名で列を特定 (部分一致, 並び順に依存しない)
    i_name = _find_col(header, "取引先名")
    i_no = _find_col(header, "取引先番号")
    i_band = _find_col(header, "距離区分")
    i_pref = _find_col(header, "都道府県")
    i_region = _find_col(header, "東西")

    def _cell(row, idx):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    customers: list[Customer] = []
    for row in rows:
        if row is None:
            continue
        name_val = _cell(row, i_name)
        no_val = _cell(row, i_no)
        # 空行スキップ (取引先名・取引先番号がともに空)
        name_raw = "" if name_val is None else str(name_val).strip()
        customer_no = "" if no_val is None else str(no_val).strip()
        if not name_raw and not customer_no:
            continue

        name_norm_base, site_token = normalize_place(name_raw)

        band_val = _cell(row, i_band)
        distance_band = "" if band_val is None else str(band_val).strip()
        km_lower, km_upper = parse_band(band_val)

        pref_val = _cell(row, i_pref)
        prefecture = None if pref_val is None or str(pref_val).strip() == "" else str(pref_val).strip()

        region_val = _cell(row, i_region)
        region = None if region_val is None or str(region_val).strip() == "" else str(region_val).strip()

        customers.append(
            Customer(
                customer_no=customer_no,
                name_raw=name_raw,
                name_norm_base=name_norm_base,
                site_token=site_token,
                distance_band=distance_band,
                km_lower=km_lower,
                km_upper=km_upper,
                prefecture=prefecture,
                region=region,
            )
        )

    return customers
