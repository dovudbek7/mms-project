"""勤怠ローダ (attendance loader).

出勤簿_日別詳細 (.xlsx) を読み, 1 行 = 1 (社員, 日付) の AttendanceDay に変換する.
データ契約 §1(c). 各ファイルは 4 シート (フレックス/固定勤務 × 一般職/管理職 等) を持ち,
ヘッダ行 0 で列を名前選択する. ファイルにより列構成 (wide/narrow) が異なるため
ヘッダ名 -> index マップを作り .get で欠落を許容する.

注意: 2 ファイルは 4月分(A) と 6月分(B) のみで, 出張月の 5月 を含まない.
在席状態 (presence) は出勤時刻・テレワーク・カレンダー・申請内容から派生する.
"""
from __future__ import annotations

import os

import openpyxl

from models import (
    AttendanceDay,
    LEAVE_CONTENTS,
    PRESENCE_HOLIDAY_WORK,
    PRESENCE_LEAVE,
    PRESENCE_NONWORK,
    PRESENCE_OFFICE,
    PRESENCE_TELEWORK,
    PRESENCE_UNKNOWN,
)
from normalize import norm, parse_excel_dt, parse_jp_date, parse_time

# 休日 (法定外/法定内) を表すカレンダー種別
_HOLIDAY_CALENDARS = {"法定外", "法定内"}


def _cell(row, idx_map, name):
    """ヘッダ名でセル値を取得. 列欠落・None は None を返す."""
    i = idx_map.get(name)
    if i is None or i >= len(row):
        return None
    v = row[i]
    if v is None or v == "":
        return None
    return v


def _movetime(row, idx_map, name):
    """移動時刻セルを time に. 実データは '2026-04-16 17:30' のような
    full datetime 文字列のため, まず parse_excel_dt で日時化し時刻部を取る.
    純粋な 'HH:MM(:SS)' 文字列にも parse_time でフォールバックする."""
    v = _cell(row, idx_map, name)
    if v is None:
        return None
    dt = parse_excel_dt(v)
    if dt is not None:
        return dt.time()
    return parse_time(v)


def _as_str(v):
    """空でない文字列なら返し, それ以外は None."""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _derive_presence(calendar_type, clock_in, telework_in,
                     holiday_work_minutes, application_content):
    """在席状態を順序付きルールで派生する (データ契約 §1(c))."""
    # 1: 出勤時刻あり -> 在席(出社)
    if clock_in is not None:
        # ただしカレンダーが休日かつ出社実態がある場合は 3 を優先判定
        if calendar_type in _HOLIDAY_CALENDARS:
            return PRESENCE_HOLIDAY_WORK
        return PRESENCE_OFFICE
    # 2: テレワーク出勤あり -> 在席(テレワーク)
    if telework_in is not None:
        return PRESENCE_TELEWORK
    # 3: 休日カレンダー かつ (出勤 or 休出時間 or 申請内容に '休日出勤')
    if calendar_type in _HOLIDAY_CALENDARS:
        has_holiday_work = (
            clock_in is not None
            or holiday_work_minutes
            or (application_content is not None and "休日出勤" in application_content)
        )
        if has_holiday_work:
            return PRESENCE_HOLIDAY_WORK
        # 4: 休日カレンダー かつ 出勤なし -> 非労働日
        return PRESENCE_NONWORK
    # 5: 平日 かつ 申請内容(カンマ分割) が休暇種別と交差 -> 休暇
    if calendar_type == "平日" and application_content:
        contents = {c.strip() for c in application_content.split(",")}
        if contents & LEAVE_CONTENTS:
            return PRESENCE_LEAVE
    # 6: それ以外 -> 不明
    return PRESENCE_UNKNOWN


def load_attendance(paths: list[str]) -> list[AttendanceDay]:
    """勤怠ファイル群を読み AttendanceDay のリストを返す.

    各パスを openpyxl(read_only=True, data_only=True) で開き, 4 シートを順に処理する.
    ヘッダ行 0 から 名前->index マップを作り, 社員番号 と 氏名 が共に空の行は skip する.
    """
    days: list[AttendanceDay] = []
    for path in paths:
        source_file = os.path.basename(path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = ws.iter_rows(values_only=True)
                try:
                    header = next(rows)
                except StopIteration:
                    continue
                idx_map = {}
                holiday_cols = []
                for i, h in enumerate(header):
                    if h is None:
                        continue
                    name = str(h).strip()
                    # 同名重複時は最初の列を採用
                    idx_map.setdefault(name, i)
                    if "休出時間" in name:
                        holiday_cols.append(i)

                for row in rows:
                    emp_raw = _cell(row, idx_map, "社員番号")
                    name_raw = _cell(row, idx_map, "氏名")
                    # 社員番号 と 氏名 が共に空ならデータ行ではない
                    if emp_raw is None and name_raw is None:
                        continue

                    calendar_type = _as_str(_cell(row, idx_map, "カレンダー"))
                    application_content = _as_str(_cell(row, idx_map, "申請内容"))

                    clock_in = parse_excel_dt(_cell(row, idx_map, "出勤時刻"))
                    clock_out = parse_excel_dt(_cell(row, idx_map, "退勤時刻"))
                    telework_in = parse_excel_dt(_cell(row, idx_map, "テレワーク出勤時刻"))
                    telework_out = parse_excel_dt(_cell(row, idx_map, "テレワーク退勤時刻"))

                    # 休出時間系のいずれかに非空値があれば True
                    holiday_work_minutes = any(
                        i < len(row) and row[i] not in (None, "")
                        for i in holiday_cols
                    )

                    presence = _derive_presence(
                        calendar_type, clock_in, telework_in,
                        holiday_work_minutes, application_content,
                    )

                    days.append(AttendanceDay(
                        emp_id_raw=str(emp_raw) if emp_raw is not None else "",
                        name_raw=str(name_raw) if name_raw is not None else "",
                        name_norm=norm(name_raw),
                        department=_as_str(_cell(row, idx_map, "部門")),
                        role=_as_str(_cell(row, idx_map, "役職")),
                        work_date=parse_jp_date(_cell(row, idx_map, "日付")),
                        weekday=_as_str(_cell(row, idx_map, "曜日")),
                        calendar_type=calendar_type,
                        application_content=application_content,
                        clock_in=clock_in,
                        clock_out=clock_out,
                        telework_in=telework_in,
                        telework_out=telework_out,
                        move_start=_movetime(row, idx_map, "移動開始"),
                        move_end=_movetime(row, idx_map, "移動終了"),
                        move2_start=_movetime(row, idx_map, "移動2開始"),
                        move2_end=_movetime(row, idx_map, "移動2終了"),
                        holiday_work_minutes=holiday_work_minutes,
                        remark=_as_str(_cell(row, idx_map, "備考")),
                        presence=presence,
                        source_file=source_file,
                        sheet_name=sheet_name,
                    ))
        finally:
            wb.close()
    return days
